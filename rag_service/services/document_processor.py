"""
Document Processor - Xử lý các loại file (docx, txt, pdf, xlsx)
Extract text và chunk thành các đoạn nhỏ
"""
import io
import re
from typing import List
from pathlib import Path
import logging

# Document processing libraries
import docx
import PyPDF2
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class DocumentChunk:
    def __init__(self, chunk_id: str, file_id: str, file_name: str, 
                 text: str, chunk_index: int, start_index: int, end_index: int):
        self.chunk_id = chunk_id
        self.file_id = file_id
        self.file_name = file_name
        self.text = text
        self.chunk_index = chunk_index
        self.start_index = start_index
        self.end_index = end_index
    
    def to_dict(self):
        return {
            'chunk_id': self.chunk_id,
            'file_id': self.file_id,
            'file_name': self.file_name,
            'text': self.text,
            'chunk_index': self.chunk_index,
            'start_index': self.start_index,
            'end_index': self.end_index
        }

class DocumentProcessor:
    CHUNK_SIZE = 500  # Kích thước mỗi chunk (số ký tự)
    CHUNK_OVERLAP = 50  # Số ký tự overlap giữa các chunk
    
    def __init__(self):
        pass
    
    async def process_document(self, file_content: bytes, file_name: str, file_id: str = None) -> List[DocumentChunk]:
        """
        Xử lý file và trả về danh sách các chunks
        """
        if file_id is None:
            import uuid
            file_id = f"DOC-{str(uuid.uuid4())[:8]}"
        
        extension = Path(file_name).suffix.lower()
        text = ""
        
        try:
            if extension == '.txt':
                text = self._extract_text_from_txt(file_content)
            elif extension == '.docx':
                text = self._extract_text_from_docx(file_content)
            elif extension == '.pdf':
                text = self._extract_text_from_pdf(file_content)
            elif extension == '.xlsx':
                text = self._extract_text_from_xlsx(file_content)
            else:
                raise ValueError(f"File type {extension} is not supported")
            
            if not text or not text.strip():
                logger.warning(f"No text extracted from file {file_name}")
                return []
            
            # Chunk text thành các đoạn nhỏ
            chunks = self._chunk_text(text, file_name, file_id)
            
            logger.info(f"Processed {file_name}: {len(chunks)} chunks created from {len(text)} characters")
            return chunks
        
        except Exception as e:
            logger.error(f"Error processing document {file_name}: {str(e)}")
            raise
    
    def _extract_text_from_txt(self, file_content: bytes) -> str:
        """Extract text từ file TXT"""
        try:
            # Thử decode với UTF-8 trước
            text = file_content.decode('utf-8')
        except UnicodeDecodeError:
            # Nếu không được, thử với encoding khác
            try:
                text = file_content.decode('latin-1')
            except:
                text = file_content.decode('utf-8', errors='ignore')
        return text
    
    def _extract_text_from_docx(self, file_content: bytes) -> str:
        """Extract text từ file DOCX"""
        doc = docx.Document(io.BytesIO(file_content))
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return '\n'.join(text_parts)
    
    def _extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text từ file PDF"""
        text_parts = []
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        for page_num, page in enumerate(pdf_reader.pages):
            try:
                text = page.extract_text()
                if text.strip():
                    text_parts.append(f"[Page {page_num + 1}]\n{text}")
            except Exception as e:
                logger.warning(f"Error extracting text from page {page_num + 1}: {str(e)}")
        return '\n\n'.join(text_parts)
    
    def _extract_text_from_xlsx(self, file_content: bytes) -> str:
        """Extract text từ file XLSX"""
        text_parts = []
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                row_text = [cell for cell in row_text if cell.strip()]
                if row_text:
                    text_parts.append(" | ".join(row_text))
        
        return '\n'.join(text_parts)
    
    def _chunk_text(self, text: str, file_name: str, file_id: str) -> List[DocumentChunk]:
        """Chunk text thành các đoạn nhỏ với overlap"""
        chunks = []
        cleaned_text = self._clean_text(text)
        
        if not cleaned_text or not cleaned_text.strip():
            return chunks
        
        start_index = 0
        chunk_index = 0
        
        while start_index < len(cleaned_text):
            end_index = min(start_index + self.CHUNK_SIZE, len(cleaned_text))
            
            # Tìm điểm cắt tốt (kết thúc câu hoặc đoạn)
            if end_index < len(cleaned_text):
                # Tìm dấu chấm câu gần nhất
                search_range = min(100, end_index - start_index)
                last_period = cleaned_text.rfind('.', start_index, end_index)
                last_newline = cleaned_text.rfind('\n', start_index, end_index)
                
                best_break = max(last_period, last_newline)
                if best_break > start_index + self.CHUNK_SIZE // 2:
                    end_index = best_break + 1
            
            chunk_text = cleaned_text[start_index:end_index].strip()
            
            if chunk_text:
                chunk_id = f"{file_id}_chunk_{chunk_index}"
                chunk = DocumentChunk(
                    chunk_id=chunk_id,
                    file_id=file_id,
                    file_name=file_name,
                    text=chunk_text,
                    chunk_index=chunk_index,
                    start_index=start_index,
                    end_index=end_index
                )
                chunks.append(chunk)
                chunk_index += 1
            
            # Di chuyển start_index với overlap
            start_index = max(start_index + 1, end_index - self.CHUNK_OVERLAP)
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """Làm sạch text: loại bỏ khoảng trắng thừa, normalize"""
        if not text:
            return ""
        
        # Loại bỏ khoảng trắng thừa
        cleaned = re.sub(r'\s+', ' ', text, flags=re.MULTILINE)
        # Loại bỏ các ký tự đặc biệt không cần thiết
        cleaned = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', cleaned)
        
        return cleaned.strip()

