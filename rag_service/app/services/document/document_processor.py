"""
Document Processor Service - Extract text and chunk documents
"""
import io
import uuid
from typing import List
from pathlib import Path
import logging

import docx
import PyPDF2
from openpyxl import load_workbook

from app.domain.document import DocumentChunk
from app.utils.text import clean_text, chunk_text
from app.core.settings import Settings

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Service for processing documents and creating chunks"""
    
    def __init__(self):
        self.chunk_size = Settings.CHUNK_SIZE
        self.chunk_overlap = Settings.CHUNK_OVERLAP
    
    async def process_document(
        self, 
        file_content: bytes, 
        file_name: str, 
        file_id: str = None
    ) -> List[DocumentChunk]:
        """
        Process file and return list of chunks
        
        Args:
            file_content: File content as bytes
            file_name: File name
            file_id: Optional file ID (will generate if not provided)
            
        Returns:
            List of DocumentChunk objects
        """
        if file_id is None:
            file_id = f"DOC-{str(uuid.uuid4())[:8]}"
        
        extension = Path(file_name).suffix.lower()
        text = ""
        
        try:
            if extension == '.txt':
                text = self._extract_text_from_txt(file_content)
            elif extension == '.docx':
                text = self._extract_text_from_docx(file_content)
            elif extension == '.pdf':
                text = self._extract_text_from_pdf(file_content)
            elif extension == '.xlsx':
                text = self._extract_text_from_xlsx(file_content)
            else:
                raise ValueError(f"File type {extension} is not supported")
            
            if not text or not text.strip():
                logger.warning(f"No text extracted from file {file_name}")
                return []
            
            # Chunk text
            chunk_tuples = chunk_text(text, self.chunk_size, self.chunk_overlap)
            
            # Convert to domain entities
            chunks = []
            for chunk_index, (chunk_content, start_index, end_index) in enumerate(chunk_tuples):
                chunk_id = f"{file_id}_chunk_{chunk_index}"
                chunk = DocumentChunk(
                    chunk_id=chunk_id,
                    file_id=file_id,
                    file_name=file_name,
                    text=chunk_content,
                    chunk_index=chunk_index,
                    start_index=start_index,
                    end_index=end_index
                )
                chunks.append(chunk)
            
            logger.info(f"Processed {file_name}: {len(chunks)} chunks created from {len(text)} characters")
            return chunks
        
        except Exception as e:
            logger.error(f"Error processing document {file_name}: {str(e)}")
            raise
    
    def _extract_text_from_txt(self, file_content: bytes) -> str:
        """Extract text from TXT file"""
        try:
            text = file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text = file_content.decode('latin-1')
            except:
                text = file_content.decode('utf-8', errors='ignore')
        return text
    
    def _extract_text_from_docx(self, file_content: bytes) -> str:
        """Extract text from DOCX file"""
        doc = docx.Document(io.BytesIO(file_content))
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return '\n'.join(text_parts)
    
    def _extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file"""
        text_parts = []
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        for page_num, page in enumerate(pdf_reader.pages):
            try:
                text = page.extract_text()
                if text.strip():
                    text_parts.append(f"[Page {page_num + 1}]\n{text}")
            except Exception as e:
                logger.warning(f"Error extracting text from page {page_num + 1}: {str(e)}")
        return '\n\n'.join(text_parts)
    
    def _extract_text_from_xlsx(self, file_content: bytes) -> str:
        """Extract text from XLSX file"""
        text_parts = []
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                row_text = [cell for cell in row_text if cell.strip()]
                if row_text:
                    text_parts.append(" | ".join(row_text))
        
        return '\n'.join(text_parts)

